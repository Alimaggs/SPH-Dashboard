#!/usr/bin/env python3
"""Build the SPH Activity Dashboard.

Reads the master list workbook, normalises it, and injects the result into
src/dashboard.template.html to produce a single self-contained HTML file.

    python scripts/build_dashboard.py

Re-run this whenever the spreadsheet changes.
"""

from __future__ import annotations

import base64
import json
from collections import Counter
import re
import sys
from functools import lru_cache
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "Data"
# Pre-sized by scripts/prepare_images.py, so this build stays deterministic.
IMAGES_DIR = ROOT / "src" / "images"
IMAGES = {
    "__SPH_ICON__":  "sph-icon.png",
    "__FAVICON__":   "favicon.png",
    "__CC_COLOUR__": "cc-colour.png",
    "__CC_WHITE__":  "cc-white.png",
}
TEMPLATE = ROOT / "src" / "dashboard.template.html"
# Sevalla serves this directory as the static site's publish directory.
PUBLIC = ROOT / "public"
OUTPUT = PUBLIC / "index.html"

# Reporting period date ranges, per the master list reference document.
PERIODS = {
    "RP1": ("RP1-26-27", "1 Sep 2026", "30 Nov 2026"),
    "RP2": ("RP2-26-27", "1 Dec 2026", "28 Feb 2027"),
    "RP3": ("RP3-26-27", "1 Mar 2027", "31 May 2027"),
    "RP4": ("RP4-26-27", "1 Jun 2027", "31 Aug 2027"),
}

# Event Name cell fill -> publishing pipeline status.
STATUS_BY_FILL = {
    "FFFFFF00": ("new", "Not yet on website"),
    "FF00B0F0": ("built", "Built, not public"),
    "FF92D050": ("live", "Live on website"),
    "FF7030A0": ("ticketing", "Needs ticketing"),
    "FFFFA500": ("later", "Later session"),
    "FFFFC000": ("later", "Later session"),
}

# Mailing List Subscriptions is free text and some values contain commas
# ("Communication, Language and Literacy") while others are comma-separated
# compounds ("Leadership, PSED"). Match greedily against the known vocabulary.
MAILING_VOCAB = [
    "Communication, Language and Literacy",
    "Leadership and Staff Development",
    "EYFS Learning Community",
    "Physical Development",
    "Working with Babies",
    "Spotlight on Twos",
    "Childminders",
    "Leadership",
    "Maths",
    "PSED",
    "SEND",
    "EDI",
    "Local",
]

# Free-text columns, so fold the expected wording to one spelling and let
# anything unrecognised through as typed.
SCOPE_ALIASES = {
    "regional": "Regional",
    "local": "Local",
}

REGISTRATION_ALIASES = {
    "website": "Website",
    "external": "External platform",
    "external platform": "External platform",
}

# One physical venue, typed several different ways across the rows.
ST_PAULS = "St Pauls Nursery Community Room"
VENUE_ALIASES = {
    "st pauls nursery community room": ST_PAULS,
    "st. paul's nursery school": ST_PAULS,
    "st paul's nursery school": ST_PAULS,
    "st pauls nursery school": ST_PAULS,
    "st paul's nursery school & children's centre": ST_PAULS,
    "st pauls nursery school & childrens centre": ST_PAULS,
    "community room": ST_PAULS,
}

# Response headers for the static host. The page is a single self-contained
# document: no scripts, styles, images or connections come from anywhere but
# itself, Google Fonts and its own data: URIs, so the policy can deny the rest
# outright. Inline script and style are unavoidable here (the whole page is one
# inline block), and neither reads anything a visitor controls.
#
# frame-ancestors 'self' blocks other sites from embedding the dashboard. To
# embed it in beyth.co.uk, add that origin here.
CSP = (
    "default-src 'none'; "
    "script-src 'unsafe-inline'; "
    "style-src 'unsafe-inline' https://fonts.googleapis.com; "
    "font-src https://fonts.gstatic.com; "
    "img-src data:; "
    "base-uri 'none'; "
    "form-action 'none'; "
    "frame-ancestors 'self'"
)

HEADERS = "\n".join([
    "/*",
    "  X-Robots-Tag: noindex, nofollow, noarchive",
    "  X-Content-Type-Options: nosniff",
    "  X-Frame-Options: SAMEORIGIN",
    "  Referrer-Policy: strict-origin-when-cross-origin",
    "  Permissions-Policy: geolocation=(), camera=(), microphone=(), payment=(), usb=()",
    "  Strict-Transport-Security: max-age=31536000",
    f"  Content-Security-Policy: {CSP}",
    "",
])


MONTHS = ["January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]


@lru_cache(maxsize=1)
def find_workbook() -> Path:
    """Locate the master list, tolerating a renamed file.

    The spreadsheet is often replaced by uploading a new one through the
    GitHub website, where the name may not match exactly.
    """
    preferred = DATA_DIR / "2026-2027 SPH Activity Master List.xlsx"
    if preferred.exists():
        return preferred

    candidates = [f for f in DATA_DIR.glob("*.xlsx") if not f.name.startswith("~$")]
    if not candidates:
        sys.exit(f"No .xlsx file found in {DATA_DIR}.")

    newest = max(candidates, key=lambda f: f.stat().st_mtime)
    if len(candidates) > 1:
        print(f"note: {len(candidates)} spreadsheets in Data/, using the newest")
    print(f"reading {newest.name}")
    return newest


def embed_image(name: str) -> str:
    """Return a prepared image as a data URI."""
    encoded = base64.b64encode((IMAGES_DIR / name).read_bytes()).decode("ascii")
    return f"data:image/png;base64,{encoded}"


def clean(value) -> str:
    """Trim a cell to a string, collapsing whitespace; None becomes ''."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def strip_tag(value: str, prefix: str) -> str:
    """Turn a HubSpot list tag into a plain label.

    'PD Communication and Language RP3-26-27' -> 'Communication and Language'
    """
    text = clean(value)
    if not text:
        return ""
    if text.startswith(prefix + " "):
        text = text[len(prefix) + 1:]
    return re.sub(r"\s*RP[1-4]-\d{2}-\d{2}$", "", text).strip()


def split_mailing(value: str) -> list[str]:
    text = clean(value)
    out: list[str] = []
    while text:
        for known in sorted(MAILING_VOCAB, key=len, reverse=True):
            if text.lower().startswith(known.lower()):
                out.append(known)
                text = text[len(known):].lstrip(", ").strip()
                break
        else:
            # Unrecognised value: keep it whole rather than mangling it.
            out.append(text)
            break
    return out


def as_bool(value) -> bool | None:
    """Read a checkbox column, whether it arrives as a bool, a word or a formula.

    'Processed for Reports' holds an Excel checkbox, which openpyxl reports as
    the formula '=FALSE()' unless the cached value is read instead.
    """
    if isinstance(value, bool):
        return value
    text = clean(value).lower().strip("=()")
    if text in ("true", "yes", "y", "1"):
        return True
    if text in ("false", "no", "n", "0"):
        return False
    return None


def title_key(name: str) -> str:
    """Normalise a title so sessions of one programme share a key."""
    stem = re.sub(r"\s*[-–]\s*Session\s*[123]\s*$", "", name, flags=re.I)
    return re.sub(r"[^a-z0-9]+", " ", stem.lower()).strip()


def parse_time(value: str) -> tuple[str, int | None]:
    """Return (display, start minutes past midnight) for sorting."""
    text = clean(value)
    match = re.match(r"(\d{1,2})(?::(\d{2}))?\s*(am|pm)", text, flags=re.I)
    if not match:
        return text, None
    hour = int(match.group(1)) % 12
    minute = int(match.group(2) or 0)
    if match.group(3).lower() == "pm":
        hour += 12
    return text, hour * 60 + minute


def read_rows() -> list[dict]:
    workbook = openpyxl.load_workbook(find_workbook())
    # A second pass with the formulas evaluated. The first pass has to keep
    # them, because it is also where the cell fill colours are read from.
    computed = openpyxl.load_workbook(find_workbook(), data_only=True)
    rows: list[dict] = []

    for sheet in workbook.worksheets:
        period = sheet.title.strip().upper()
        headers = [clean(c.value) for c in next(sheet.iter_rows(min_row=1, max_row=1))]
        index = {h: i for i, h in enumerate(headers)}
        values_sheet = computed[sheet.title]

        def cell(row, header):
            position = index.get(header)
            return row[position] if position is not None else None

        def cell_value(row, header):
            """The evaluated result for a formula cell."""
            position = index.get(header)
            if position is None:
                return None
            return values_sheet.cell(row=row[0].row, column=position + 1).value

        for excel_row in sheet.iter_rows(min_row=2):
            name_cell = cell(excel_row, "Event Name")
            name = clean(name_cell.value if name_cell else None)
            if not name:
                continue

            fill = name_cell.fill
            rgb = None
            if fill is not None and fill.fill_type == "solid" and fill.fgColor.type == "rgb":
                rgb = fill.fgColor.rgb
            status, status_label = STATUS_BY_FILL.get(rgb, ("unknown", "Not set"))

            raw_date = cell(excel_row, "Activity Date").value
            if isinstance(raw_date, datetime):
                activity = raw_date.date()
            elif isinstance(raw_date, date):
                activity = raw_date
            else:
                activity = None

            session_match = re.search(r"[-–]\s*Session\s*([123])\s*$", name, flags=re.I)
            session = int(session_match.group(1)) if session_match else None

            location = clean(cell(excel_row, "Location").value)
            venue = VENUE_ALIASES.get(location.lower(), location)
            if location.lower() == "online":
                fmt, venue = "Online", ""
            elif location.upper() == "TBC" or not location:
                fmt, venue = "Venue TBC", ""
            else:
                fmt = "Face-to-face"

            # The column usually holds a capacity, but reads "External" where
            # booking happens elsewhere and the number is not ours to know.
            # Keep that word rather than showing the row as missing data.
            tickets_raw = cell(excel_row, "Number of Tickets").value
            try:
                tickets = int(tickets_raw)
                tickets_text = ""
            except (TypeError, ValueError):
                tickets = None
                tickets_text = clean(tickets_raw)

            scope_cell = cell(excel_row, "Regional/Local")
            scope = clean(scope_cell.value if scope_cell else None)
            registration_cell = cell(excel_row, "Registration")
            registration = clean(registration_cell.value if registration_cell else None)

            time_display, time_sort = parse_time(cell(excel_row, "Time").value)
            brochure = clean(cell(excel_row, "Brochure Category").value)
            # 'Working With Babies' and 'Working with Babies' are one group.
            if brochure.lower() == "working with babies":
                brochure = "Working with Babies"

            rows.append({
                "rp": period,
                "name": name,
                "stem": re.sub(r"\s*[-–]\s*Session\s*[123]\s*$", "", name, flags=re.I),
                "key": title_key(name),
                "session": session,
                "date": activity.isoformat() if activity else "",
                # Day of Week is a formula cell, so derive the label from the date.
                "day": activity.strftime("%A") if activity else "",
                "month": f"{activity.year}-{activity.month:02d}" if activity else "",
                "monthLabel": f"{MONTHS[activity.month - 1]} {activity.year}" if activity else "Date TBC",
                "time": time_display,
                "timeSort": time_sort if time_sort is not None else 9999,
                "brochure": brochure,
                "location": location,
                "venue": venue,
                "format": fmt,
                "tickets": tickets,
                "ticketsText": tickets_text,
                "type": strip_tag(cell(excel_row, "Activity Type").value, "AT"),
                "pd": strip_tag(cell(excel_row, "Professional Development Category").value, "PD"),
                "network": strip_tag(cell(excel_row, "Network").value, "NW"),
                "mailing": split_mailing(cell(excel_row, "Mailing List Subscriptions").value),
                "scope": SCOPE_ALIASES.get(scope.lower(), scope),
                "registration": REGISTRATION_ALIASES.get(registration.lower(), registration),
                "processed": as_bool(cell_value(excel_row, "Processed for Reports")),
                "cpd": clean(cell(excel_row, "CPD Bundle for Survey").value),
                "workflows": clean(cell(excel_row, "Workflows Configured").value),
                "notes": clean(cell(excel_row, "Notes").value),
                "url": clean(cell(excel_row, "Website URL").value),
                "summary": clean(cell(excel_row, "Short Description").value),
                "status": status,
                "statusLabel": status_label,
                "tags": {
                    "registered": clean(cell(excel_row, "Period Registered").value),
                    "attended": clean(cell(excel_row, "Period Attended").value),
                    "type": clean(cell(excel_row, "Activity Type").value),
                    "pd": clean(cell(excel_row, "Professional Development Category").value),
                    "network": clean(cell(excel_row, "Network").value),
                },
            })

    rows.sort(key=lambda r: (r["date"] or "9999", r["timeSort"], r["name"]))

    # Mark multi-part programmes: a title stem carrying "Session N" rows.
    parts: dict[str, int] = {}
    for row in rows:
        if row["session"]:
            parts[row["key"]] = max(parts.get(row["key"], 0), row["session"])
    for i, row in enumerate(rows, start=1):
        row["id"] = f"a{i:03d}"
        row["parts"] = parts.get(row["key"], 0)
        # A "distinct activity" is every row except sessions 2 and 3 of a
        # multi-part programme, which are covered by the session 1 booking.
        row["counts"] = not (row["session"] and row["session"] > 1)
    return rows


def build() -> None:
    rows = read_rows()
    payload = {
        "generated": datetime.now().strftime("%d %B %Y"),
        "year": datetime.now().year,
        "periods": [
            {"id": key, "tag": tag, "from": start, "to": end}
            for key, (tag, start, end) in PERIODS.items()
        ],
        "activities": rows,
    }

    html = TEMPLATE.read_text(encoding="utf-8")
    for placeholder, name in IMAGES.items():
        html = html.replace(placeholder, embed_image(name))
    html = html.replace(
        '"__SPH_DATA__"',
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
    )
    if "__SPH_" in html:
        sys.exit("Template still contains an unreplaced placeholder.")

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(html, encoding="utf-8")

    # Keep the dashboard out of search results. Three overlapping signals,
    # because each can be ignored independently: the noindex meta tag in the
    # page, robots.txt, and an X-Robots-Tag header via Sevalla's _headers file.
    (PUBLIC / "robots.txt").write_text(
        "User-agent: *\nDisallow: /\n", encoding="utf-8")
    (PUBLIC / "_headers").write_text(HEADERS, encoding="utf-8")

    counted = sum(1 for r in rows if r["counts"])
    size = OUTPUT.stat().st_size / 1024
    print(f"{len(rows)} rows -> {counted} distinct activities")
    for field, label in (("scope", "Regional/Local"),
                         ("registration", "Registration"),
                         ("processed", "Processed for Reports")):
        tally = Counter(r[field] if r[field] != "" else "(blank)" for r in rows)
        print(f"  {label}: " + ", ".join(f"{v}={n}" for v, n in tally.most_common()))
    print(f"wrote {OUTPUT.relative_to(ROOT)} ({size:.0f} KB)")


if __name__ == "__main__":
    build()
